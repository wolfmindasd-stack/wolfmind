"""Excel export utility for Wolf's Mind gestionale."""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def _write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for row in rows:
        ws.append(row)
    # Auto-width
    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = min(max(max_len + 2, 10), 60)


def generate_backup_xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # Tesserati
    _write_sheet(wb, "Tesserati",
        ["N. Tessera", "Cognome", "Nome", "CF", "Data nascita", "Indirizzo",
         "CAP", "Città", "Provincia", "Email", "Telefono",
         "Scad. tesseramento", "Scad. visita medica", "Note"],
        [[t.get("numero_tessera", ""), t.get("cognome", ""), t.get("nome", ""),
          t.get("codice_fiscale", ""), t.get("data_nascita", ""),
          f"{t.get('indirizzo', '')} {t.get('civico', '')}".strip(),
          t.get("cap", ""), t.get("citta", ""), t.get("provincia", ""),
          t.get("email", ""), t.get("telefono", ""),
          t.get("scadenza_tesseramento", ""), t.get("scadenza_visita_medica", ""),
          t.get("note", "")] for t in data.get("tesserati", [])])

    # Ricevute
    _write_sheet(wb, "Ricevute",
        ["Numero", "Data", "Tesserato", "Metodo pagamento", "Totale",
         "Emessa per", "Annullata", "Note"],
        [[r.get("numero", ""), r.get("data", "")[:10], r.get("tesserato_nome", ""),
          r.get("metodo_pagamento", ""), float(r.get("totale") or 0),
          r.get("emesso_per_nome") or r.get("emesso_da_nome", ""),
          "Sì" if r.get("annullata") else "No", r.get("note", "")]
         for r in data.get("ricevute", [])])

    # Voci ricevute (dettaglio)
    voci_rows = []
    for r in data.get("ricevute", []):
        for it in r.get("items", []):
            voci_rows.append([r.get("numero", ""), r.get("data", "")[:10],
                              r.get("tesserato_nome", ""), it.get("descrizione", ""),
                              it.get("num_lezioni") or "", float(it.get("importo") or 0),
                              "Sì" if it.get("esclude_da_compensi") else "No"])
    _write_sheet(wb, "Ricevute-Dettaglio",
        ["Numero", "Data", "Tesserato", "Descrizione", "N. lezioni", "Importo",
         "Escluso compensi"], voci_rows)

    # Movimenti
    _write_sheet(wb, "Movimenti",
        ["Data", "Tipo", "Categoria", "Descrizione", "Importo", "Tecnico"],
        [[m.get("data", "")[:10], m.get("tipo", ""), m.get("categoria", ""),
          m.get("descrizione", ""), float(m.get("importo") or 0),
          m.get("tecnico_nome") or ""] for m in data.get("movimenti", [])])

    # Abbonamenti
    _write_sheet(wb, "Abbonamenti",
        ["Tesserato", "Descrizione", "Data acquisto", "N. lezioni",
         "Effettuate", "Residue", "Prezzo"],
        [[a.get("tesserato_nome", ""), a.get("descrizione", ""),
          a.get("data_acquisto", "")[:10], a.get("num_lezioni_totali") or "",
          a.get("lezioni_effettuate", 0), a.get("lezioni_residue") or "",
          float(a.get("prezzo") or 0)] for a in data.get("abbonamenti", [])])

    # Lezioni
    _write_sheet(wb, "Lezioni",
        ["Data", "Luogo", "Tecnico", "N. Partecipanti", "Partecipanti", "Note"],
        [[l.get("data", "")[:10], l.get("luogo", ""), l.get("tecnico_nome") or "",
          len(l.get("partecipanti", [])),
          ", ".join(p.get("nome_completo", "") for p in l.get("partecipanti", [])),
          l.get("note", "")] for l in data.get("lezioni", [])])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
